from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.models import User

from .models import *
from .forms import *
from django.core.files import File
import os.path
import os

from openpyxl import *

def uvozi_s_siframi(imedatoteke, imeskupine):
    print('\x1b[6;30;42m' + 'Začetek uvažanja '+ imedatoteke + '\x1b[0m')

    #dodajanje skupin
    skupina_izdelkov = None
    if SkupinaIzdelkov.objects.filter(ime=imeskupine).exists():
        skupina_izdelkov = SkupinaIzdelkov.objects.get(ime=imeskupine)
    else:
        skupina_izdelkov = SkupinaIzdelkov(ime= imeskupine)
        skupina_izdelkov.save()


    wb = load_workbook('products/'+ imedatoteke + '.xlsx')
    sheet = wb.active
    
    #cellsa = sheet['A3':'F251']
    for row in range(5,sheet.max_row):
        if(sheet.cell(row,3).value is None):
            break
        
        sifra = sheet.cell(row,2).value
        #artikel = sheet.cell(row,3).value
        ean_koda = sheet.cell(row,5).value
        opis = sheet.cell(row,4).value
        opis_dimenzij = sheet.cell(row,6).value
        opis = opis + " "+ opis_dimenzij
        tag = sheet.cell(row,7).value
        tag = tag.split(", ")

        #odstrani empty stringe ce se kdo ponesreci zmoti
        list_tag = list(filter(None,tag))
        #print(list_tag)

        #dodajanje tagov
        object_list_tag = []
        for en_tag in list_tag:
            if Tag.objects.filter(ime=en_tag).exists():
                object_list_tag.append(Tag.objects.get(ime=en_tag))
            else:
                new_tag = Tag(ime=en_tag)
                new_tag.save()
                object_list_tag.append(new_tag)

        


        #dodajanje izdelkov

        new_izdelek = Izdelek(ime=sifra,opis=opis,skupina_izdelkov=skupina_izdelkov,koda=sifra, ean_koda=ean_koda)
        pot_do_slike = "products/slike/"+sifra+".jpg"

        try:
            new_izdelek.slika.save(sifra+ ".jpg", File(open(pot_do_slike,'rb')))
        except IOError:
            new_izdelek.slika.save("logo.jpg", File(open("products/slike/logo.jpg",'rb')))
            #print(sifra)

        new_izdelek.save()
        new_izdelek.tag.add(*object_list_tag)
        new_izdelek.save()
            

            

        
    print('\x1b[6;30;42m' + 'Konec uvažanja izdelkov ' + imedatoteke+ '\x1b[0m')
        
    return

def uvozi_brez_siframi(imedatoteke, imeskupine, sheet_name):
    print('\x1b[6;30;42m' + 'Začetek uvažanja '+ imedatoteke + '\x1b[0m')

    #dodajanje skupin
    skupina_izdelkov = None
    if SkupinaIzdelkov.objects.filter(ime=imeskupine).exists():
        skupina_izdelkov = SkupinaIzdelkov.objects.get(ime=imeskupine)
    else:
        skupina_izdelkov = SkupinaIzdelkov(ime= imeskupine)
        skupina_izdelkov.save()


    wb = load_workbook('products/'+ imedatoteke + '.xlsx')
    sheet = wb[sheet_name]
    
    #cellsa = sheet['A3':'F251']
    for row in range(5,sheet.max_row):
        if(sheet.cell(row,3).value is None):
            break
        
        sifra = sheet.cell(row,5).value
        artikel = sheet.cell(row,3).value
        ean_koda = sheet.cell(row,5).value
        opis = sheet.cell(row,4).value
        opis_dimenzij = sheet.cell(row,6).value
        opis = opis + " "+ opis_dimenzij
        tag = sheet.cell(row,7).value
        tag = tag.split(", ")

        #odstrani empty stringe ce se kdo ponesreci zmoti
        list_tag = list(filter(None,tag))
        #print(list_tag)

        #dodajanje tagov
        object_list_tag = []
        for en_tag in list_tag:
            if Tag.objects.filter(ime=en_tag).exists():
                object_list_tag.append(Tag.objects.get(ime=en_tag))
            else:
                new_tag = Tag(ime=en_tag)
                new_tag.save()
                object_list_tag.append(new_tag)

        


        #dodajanje izdelkov

        new_izdelek = Izdelek(ime=artikel,opis=opis,skupina_izdelkov=skupina_izdelkov,koda=ean_koda, ean_koda=ean_koda)
        pot_do_slike = "products/slike/nislike.jpg"

        try:
            new_izdelek.slika.save("nislike.jpg", File(open(pot_do_slike,'rb')))
        except IOError:
            new_izdelek.slika.save("logo.jpg", File(open("products/slike/logo.jpg",'rb')))
            #print(sifra)

        new_izdelek.save()
        new_izdelek.tag.add(*object_list_tag)
        new_izdelek.save()
            

            

        
    print('\x1b[6;30;42m' + 'Konec uvažanja izdelkov ' + imedatoteke+ '\x1b[0m')
        
    return

def uvozi_izdelke():
    uvozi_s_siframi("Seznam_artiklov_Sifrant_razglednice_ok", "Razglednice")
    uvozi_s_siframi("Seznam_artiklov_Sifrant_magnet_kovinski_ok", "Magneti")
    uvozi_brez_siframi("Seznam_artiklov_Sifrant_knjige_zemljevidi_sid_ok", "Knjige","knjige")
    uvozi_brez_siframi("Seznam_artiklov_Sifrant_knjige_zemljevidi_sid_ok", "Zemljevidi","zemljevidi")
    


def naredi_bazo(request):

    
    #naredi potnika
    user, created = User.objects.get_or_create(username="potnik", email="potnik@gmail.com")
    user.first_name = "Jure"
    user.last_name = "Hovelja"

    if created:
        user.set_password("adminadmin")
        user.is_staff=False
        user.is_superuser=False
    
    user.save()

    #potnik = Potnik(user = user, prodajno_mesto = prodajno_mesto, telefon = "031786453", email = "potnik@gmail.com")
    #potnik.save()

    potnik_uporabnik = Uporabnik(user = user, je_potnik = True)
    potnik_uporabnik.save()

    potnik = Potnik(uporabnik = potnik_uporabnik,  telefon = "031786453", email = "potnik@gmail.com")
    potnik.save()



    prodajno_mesto = ProdajnoMesto(ime = "Trafika Škofja Loka", naslov = "Škofja Loka 88", postna_stevilka = "4220", obcina = "Škofja Loka", kontaktna_oseba = "Anja Novak", telefon = "041500677", potnik=potnik)
    prodajno_mesto.save()

    prodajno_mesto1 = ProdajnoMesto(ime = "Trafika Ljubljana", naslov = "Cankarjeva ulica 1", postna_stevilka = "1000", obcina = "Ljubljana", kontaktna_oseba = "Luka Cvek", telefon = "041456289", potnik=potnik)
    prodajno_mesto1.save()

    podjetje = Podjetje(ime = "trafika d.o.o.", naslov = "Škofja Loka 12", postna_stevilka = "4220", obcina = "Škofja Loka", davcna_stevilka = "67294308")
    podjetje.save()

    #uporabnik1 = Uporabnik(user = User.objects.get(username='admin'), podjetje = podjetje, prodajno_mesto = prodajno_mesto)
    #uporabnik1.save()
    #naredi uporabnika
    user_uporabnik, created = User.objects.get_or_create(username="uporabnik2", email="uporabnik2@gmail.com")
    user_uporabnik.first_name = "Kristina"
    user_uporabnik.last_name = "Listnik"

    if created:
        user_uporabnik.set_password("adminadmin")
        user_uporabnik.is_staff=False
        user_uporabnik.is_superuser=False

    user_uporabnik.save()
    uporabnik = Uporabnik(user = user_uporabnik, podjetje = podjetje, prodajno_mesto = prodajno_mesto)
    uporabnik.save()
    
    #naredi uporabnika
    user_uporabnik, created = User.objects.get_or_create(username="uporabnik", email="uporabnik@gmail.com")
    user_uporabnik.first_name = "Mitja"
    user_uporabnik.last_name = "Resnik"

    if created:
        user_uporabnik.set_password("adminadmin")
        user_uporabnik.is_staff=False
        user_uporabnik.is_superuser=False

    user_uporabnik.save()
    uporabnik = Uporabnik(user = user_uporabnik, podjetje = podjetje, prodajno_mesto = prodajno_mesto)
    uporabnik.save()

    
    uvozi_izdelke()
    

    ######## UPORABLJAMO SEDAJ REALNE IZDELKE
    '''

    #
  

    skupina_izdelkov1 = SkupinaIzdelkov(ime = "Magneti", koda = "2389472983")
    skupina_izdelkov1.save()

    skupina_izdelkov2 = SkupinaIzdelkov(ime = "Razglednice", koda = "2734242398")
    skupina_izdelkov2.save()

    tag_bled = Tag(ime = "Bled")
    tag_bled.save()
    tag_slovenia = Tag(ime = "Slovenija")
    tag_slovenia.save()
    tag_bohinj = Tag(ime = "Bohinj")
    tag_bohinj.save()
    tag_piran = Tag(ime = "Piran")
    tag_piran.save()
    tag_triglav = Tag(ime = "Triglav")
    tag_triglav.save()
    tag_jama= Tag(ime = "Postojnska jama")
    tag_jama.save()
    tag_lj= Tag(ime = "Ljubljana")
    tag_lj.save()

    #Bled
    izdelek_bled = Izdelek(ime = "Magnet Bled", opis = "Dimenzija: 60x80mm", skupina_izdelkov = skupina_izdelkov1, koda = "MAG001")
    izdelek_bled.slika.save('Bled.jpg', File(open(r'media/gallery/Bled.jpg','rb')))
    izdelek_bled.tag.add(tag_slovenia,tag_bled);

    #narocilo_bled = NarociloIzdelka(izdelek = izdelek_bled,kolicina = 25)
    #narocilo_bled.save()

    izdelek_triglav = Izdelek(ime = "Magnet Triglav", opis = "Dimenzija: 60x80mm", skupina_izdelkov = skupina_izdelkov1, koda = "MAG015")
    izdelek_triglav.slika.save("Triglav.jpg", File(open(r'media/gallery/Triglav.jpg','rb')))
    izdelek_triglav.tag.add(tag_slovenia,tag_triglav);
    #narocilo_triglav = NarociloIzdelka(izdelek = izdelek_triglav,kolicina = 100)
    #narocilo_triglav.save()

    

    izdelek_razglednica1 = Izdelek(ime = "Razglednica Postojnska jama", opis = "Dimenzija: 60x80mm", skupina_izdelkov = skupina_izdelkov2, koda = "RAZ001")
    izdelek_razglednica1.slika.save("jama.jpg", File(open(r'media/gallery/jama.jpg','rb')))
    izdelek_razglednica1.tag.add(tag_slovenia,tag_jama)
    #narocilo_izdelek_razglednica1 = NarociloIzdelka(izdelek = izdelek_razglednica1,kolicina = 100)
    #narocilo_izdelek_razglednica1.save()

    #izdelek_razglednica2 = Izdelek(ime = "Razglednica test 2", opis = "Dimenzija: 60x80mm", skupina_izdelkov = skupina_izdelkov2, koda = "sadfasdf")
    #izdelek_razglednica2.slika.save("marija.jpg", File(open(r'media/gallery/magneti_zbirno_201727.jpg','rb')))
    #narocilo_izdelek_razglednica2 = NarociloIzdelka(izdelek = izdelek_razglednica2,kolicina = 100)
    #narocilo_izdelek_razglednica2.save()

    #Bohinj
    izdelek_triglav = Izdelek(ime = "Magnet Bohinj", opis = "Dimenzija: 60x80mm", skupina_izdelkov = skupina_izdelkov1, koda = "MAG003")
    izdelek_triglav.slika.save("Bohinj.jpg", File(open(r'media/gallery/bohinj.jpg','rb')))
    izdelek_triglav.tag.add(tag_slovenia,tag_bohinj);

    #Ljubljana
    izdelek_lj1 = Izdelek(ime = "Magnet Ljubljana 1", opis = "Dimenzija: 60x80mm", skupina_izdelkov = skupina_izdelkov1, koda = "MAG004")
    izdelek_lj1.slika.save("ljubljana.jpg", File(open(r'media/gallery/ljubljana.jpg','rb')))
    izdelek_lj1.tag.add(tag_slovenia,tag_lj);
    #Ljubljana1
    izdelek_lj2 = Izdelek(ime = "Magnet Ljubljana 2", opis = "Dimenzija: 60x80mm", skupina_izdelkov = skupina_izdelkov1, koda = "MAG005")
    izdelek_lj2.slika.save("ljubljana1.jpg", File(open(r'media/gallery/ljubljana1.jpg','rb')))
    izdelek_lj2.tag.add(tag_slovenia,tag_lj);
    #Ljubljana2
    izdelek_lj3 = Izdelek(ime = "Magnet Ljubljana 3", opis = "Dimenzija: 60x80mm", skupina_izdelkov = skupina_izdelkov1, koda = "MAG006")
    izdelek_lj3.slika.save("ljubljana2.jpg", File(open(r'media/gallery/ljubljana2.jpg','rb')))
    izdelek_lj3.tag.add(tag_slovenia,tag_lj);
    
    #Piran
    izdelek_piran = Izdelek(ime = "Magnet Piran 1", opis = "Dimenzija: 60x80mm", skupina_izdelkov = skupina_izdelkov1, koda = "MAG007")
    izdelek_piran.slika.save("piran.jpg", File(open(r'media/gallery/piran.jpg','rb')))
    izdelek_piran.tag.add(tag_slovenia,tag_piran);
    #Piran1
    izdelek_piran1 = Izdelek(ime = "Magnet Piran 2 ", opis = "Dimenzija: 60x80mm", skupina_izdelkov = skupina_izdelkov1, koda = "MAG008")
    izdelek_piran1.slika.save("piran1.jpg", File(open(r'media/gallery/piran1.jpg','rb')))
    izdelek_piran1.tag.add(tag_slovenia,tag_piran);
    
    #kosarica = Kosarica(uporabnik = uporabnik)
    #kosarica.save()
    #kosarica.narocila_izdelka.add(narocilo_bled,narocilo_triglav,narocilo_izdelek_razglednica1,narocilo_izdelek_razglednica2)
    #kosarica.save()
    '''
